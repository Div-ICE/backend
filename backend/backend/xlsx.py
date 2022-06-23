from openpyxl import load_workbook


def read_xlsx(file, sheet):
    """Функция принимает файл *.xlsx и название листа.
    Возвращает объект лист и его размеры"""
    file = load_workbook(file)
    sheet = file[sheet]
    rows = sheet.max_row
    cols = sheet.max_column
    file.close
    return sheet, rows, cols


def create_list(sheet, rows, cols):
    """Функция принимает объект лист и его размеры.
    Возвращает список словарей"""

    list1 = []
    for i in range(2, rows + 1):
        list = []
        dictionary = {}
        for j in range(1, cols + 1):
            cell = sheet.cell(row=i, column=j)
            if cell.value:
                header = sheet.cell(row=1, column=j)
                if header.value == '№':
                    header.value = 'number'
                list.append(cell.value)
                dictionary[header.value] = cell.value
        list1.append(dictionary)
    return list1


if __name__ == '__main__':

    file = 'client_org.xlsx'
    file1 = 'bills.xlsx'
    sheet1 = 'client'
    sheet2 = 'organization'
    sheet3 = 'Лист1'
    sheet, rows, cols = read_xlsx(file, sheet1)
    list = create_list(sheet, rows, cols)
    print(list)
